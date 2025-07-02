import os
import re
import json
import time
import glob
import random
import logging
import traceback
import requests
from io import BytesIO
from datetime import datetime
from datetime import timedelta
from collections import Counter
import pandas as pd
import numpy as np
import openpyxl
import pytesseract
import difflib
import pyautogui
import subprocess
import shutil
from bs4 import BeautifulSoup
from mss import mss
from PIL import ImageFilter, ImageStat
from collections import deque
from PIL import Image, ImageDraw, ImageEnhance, ImageOps, ImageFilter
from scipy.ndimage import label, find_objects
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from selenium import webdriver
from mcap_scraper_utils import run_detailed_mcap_scraper
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import JavascriptException
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    ElementNotInteractableException,
    StaleElementReferenceException
)

# Настройка логгера
logging.basicConfig(
    filename='gmgn_scraper.log',
    filemode='a',
    format='[%(asctime)s] [%(levelname)s] %(message)s',
    level=logging.INFO
)
# Настройки
WAIT_TIMEOUT = 2  # Увеличили время ожидания
DELAY = 5  # Уменьшили задержку между действиями
MAX_RETRIES = 2  # Увеличили количество попыток

# Пути к папкам
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DOWNLOAD_FOLDER = os.path.join(SCRIPT_DIR, 'downloads')
REPORT_FOLDER = os.path.join(SCRIPT_DIR, 'reports')
RESUME_PATH = "resume_state.json"

os.makedirs(DOWNLOAD_FOLDER, exist_ok=True)
os.makedirs(REPORT_FOLDER, exist_ok=True)

# ====== Имитация поведения человека ======

def safe_click(driver, element, timeout=10):
    """
    Аккуратный клик по элементу:
    - Дождаться видимости и кликабельности
    - Мягко навести мышку
    - Сделать небольшую паузу
    - Кликнуть
    - После клика снова пауза
    """
    try:
        WebDriverWait(driver, timeout).until(EC.visibility_of(element))
        WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((By.XPATH, element.get_attribute("xpath") or ".")))

        actions = ActionChains(driver)
        actions.move_to_element(element).pause(random.uniform(0.6, 1.3)).click().perform()

        print(f"[SAFE_CLICK] ✅ Клик выполнен безопасно по элементу: {element.tag_name}")
        time.sleep(random.uniform(1.0, 2.0))

    except Exception as e:
        print(f"[SAFE_CLICK][ERROR] ❗ Ошибка безопасного клика: {e}")

def random_scroll(driver, max_offset=200):
    """Случайный скролл вверх или вниз."""
    offset = random.randint(-max_offset, max_offset)
    driver.execute_script(f"window.scrollBy(0, {offset});")
    time.sleep(random.uniform(0.3, 0.7))

def random_mouse_move(driver, max_offset=50):
    """Имитация случайного движения мыши."""
    actions = ActionChains(driver)
    x_offset = random.randint(-max_offset, max_offset)
    y_offset = random.randint(-max_offset, max_offset)
    actions.move_by_offset(x_offset, y_offset).perform()
    time.sleep(random.uniform(0.2, 0.5))
    actions.move_by_offset(-x_offset, -y_offset).perform()
    time.sleep(random.uniform(0.2, 0.5))

def slow_typing(element, text, min_delay=0.05, max_delay=0.15):
    """Медленный посимвольный ввод текста."""
    for char in text:
        element.send_keys(char)
        time.sleep(random.uniform(min_delay, max_delay))

def save_resume_state(wallet_index=0, token_index=0, stage="start"):
    state = {
        "wallet_index": wallet_index,
        "token_index": token_index,
        "stage": stage
    }
    with open(RESUME_PATH, "w") as f:
        json.dump(state, f)
    print(f"[RESUME]  Состояние сохранено: {state}")
    logging.info(f"[RESUME] Состояние сохранено: {state}")

def load_resume_state():
    if os.path.exists(RESUME_PATH):
        try:
            with open(RESUME_PATH, "r") as f:
                try:
                    state = json.load(f)
                    print(f"[RESUME]  Загружаем состояние: {state}")
                    logging.info(f"[RESUME] Загружаем состояние: {state}")
                    return state
                except json.JSONDecodeError:
                    print("[RESUME] ❌ Повреждён файл состояния — начинаем с нуля")
                    logging.warning("[RESUME] Повреждён файл состояния — начинаем с нуля")
        except Exception as e:
            print(f"[RESUME] ⚠ Ошибка при чтении состояния: {e}")
            logging.warning(f"[RESUME] Ошибка при чтении состояния: {e}")
    return {"wallet_index": 0, "token_index": 0, "stage": "start"}

def clear_resume_state():
    """Удаляет файл с состоянием после успешной обработки кошелька."""
    if os.path.exists(RESUME_PATH):
        try:
            os.remove(RESUME_PATH)
            print("[RESUME] ✅ Файл состояния удален.")
            logging.info("[RESUME] ✅ Файл состояния удален.")
        except Exception as e:
            print(f"[RESUME] ❌ Ошибка при удалении файла состояния: {e}")

def update_resume_state(wallet_index=0, token_index=0, stage="start"):
    save_resume_state(wallet_index, token_index, stage)

def close_pnl_modal_if_present(driver):
    """Закрывает модалку 'PnL calculation in progress', если она появляется."""
    try:
        print("[MODAL] Проверка на наличие окна PnL...")
        modal = WebDriverWait(driver, 3).until(
            EC.presence_of_element_located((By.XPATH, "//div[contains(text(), 'PnL calculation in progress')]"))
        )
        print("[MODAL] ⚠ Найдено окно ожидания расчёта PnL")
        btn = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'I Know')]"))
        )
        ActionChains(driver).move_to_element(btn).pause(0.5).click().perform()
        print("[MODAL] ✅ Кнопка 'I Know' нажата")
        time.sleep(1.5)
    except TimeoutException:
        print("[MODAL] ❌ Модальное окно PnL не обнаружено (всё ок)")

def solve_turnstile_if_present(driver, max_attempts=3):
    screenshot_dir = "/tmp/turnstile_watch"

    def ensure_screenshot_dir():
        if os.path.exists(screenshot_dir):
            shutil.rmtree(screenshot_dir)
        os.makedirs(screenshot_dir)

    def capture_screenshots_periodically(interval=1.0, max_duration=30):
        ensure_screenshot_dir()
        start_time = time.time()
        print("[WATCHER]  Мониторим появление капчи (пустой квадрат)...")

        # Координаты центра чекбокса
        checkbox_x, checkbox_y = 851, 401
        
        # Расширенный регион: 500x250, сдвигаем правее и выше
        region_width, region_height = 500, 250
        offset_x = 60   # сдвиг вправо
        offset_y = 60   # сдвиг вверх
        
        left   = checkbox_x - region_width // 2 + offset_x
        top    = checkbox_y - region_height // 2 - offset_y
        right  = checkbox_x + region_width // 2 + offset_x
        bottom = checkbox_y + region_height // 2 - offset_y
        crop_box = (left, top, right, bottom)

        while time.time() - start_time < max_duration:
            timestamp = datetime.now().strftime("%H%M%S")
            path = os.path.join(screenshot_dir, f"cap_{timestamp}.png")
            driver.save_screenshot(path)
            print(f"[WATCHER]   Скриншот сохранён: {path}")

            image = Image.open(path)
            cropped = image.crop(crop_box)
            cropped = image.crop(crop_box).convert("L")
            cropped = cropped.resize((cropped.width * 3, cropped.height * 3))  # Больше увеличение
            cropped = cropped.point(lambda x: 0 if x < 160 else 255)  # Жёсткая бинаризация

            debug_crop_path = os.path.join(screenshot_dir, f"crop_{timestamp}.png")
            cropped.save(debug_crop_path)
            print(f"[DEBUG] Вырезанная область сохранена: {debug_crop_path}")
            
            text = pytesseract.image_to_string(cropped, lang="rus+eng")
            print(f"[OCR] Распознанный текст:\n{text.strip()}")

            if re.search(r'gmg[nm]\.ai', text, re.IGNORECASE):
                print("[WATCHER] ✅ Найдено совпадение с gmgn.ai — капча на экране")
                return True

            time.sleep(interval)

        print("[WATCHER]  Время вышло, квадрат не найден")
        return False

    def is_captcha_cleared():
        try:
            driver.find_element(By.XPATH, "//*[contains(text(), 'Подтвердите') or contains(text(), 'Confirm')]")
            return False
        except:
            return True

    print("[CAPTCHA] Поиск Cloudflare Turnstile...")

    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Подтвердите') or contains(text(), 'Confirm')]"))
        )
    except:
        print("[CAPTCHA] ❌ Капча не обнаружена")
        return False

    print("[CAPTCHA] ✅ Капча найдена — начинаем попытки")

    for attempt in range(1, max_attempts + 1):
        print(f"[CAPTCHA]  Попытка {attempt}/{max_attempts}")

        print("[ACTION]  Обновляем страницу (F5)...")
        driver.refresh()
        time.sleep(5)

        print("[WATCH]  Ожидаем чекбокс через OCR...")
        found = capture_screenshots_periodically(interval=1.0, max_duration=30)
        if not found:
            print("[CAPTCHA] ❌ Чекбокс не появился — пробуем снова")
            continue

        # ✅ Клик по координатам (точка центра чекбокса)
        x, y = 851, 401
        print(f"[CLICK]  Кликаем по координатам: ({x}, {y})")
        pyautogui.moveTo(x, y, duration=0.5)
        pyautogui.click()
        time.sleep(5)

        print("[WATCH] Проверка исчезновения капчи...")
        if is_captcha_cleared():
            print("[CAPTCHA] ✅ Капча успешно решена")
            shutil.rmtree(screenshot_dir, ignore_errors=True)
            return True
        else:
            print("[CAPTCHA] ⚠️ Капча всё ещё активна — следующая попытка")

    print("[CAPTCHA] ❌ Все попытки исчерпаны — не удалось решить капчу")
    return False
    
def safe_find_element(driver, by, value, retries=MAX_RETRIES, delay=DELAY):
    """Безопасное получение элемента с повторными попытками"""
    for i in range(retries):
        try:
            element = WebDriverWait(driver, WAIT_TIMEOUT).until(
                EC.presence_of_element_located((by, value))
            )
            return element
        except Exception as e:
#            print(f"[RETRY {i+1}/{retries}] Не удалось найти элемент {by}={value}: {str(e)}")
            if i == retries - 1:
                raise
            time.sleep(delay)

def safe_click(driver, element, retries=MAX_RETRIES, delay=DELAY):
    """Безопасный клик с повторными попытками"""
    for i in range(retries):
        try:
            WebDriverWait(driver, WAIT_TIMEOUT).until(
                EC.element_to_be_clickable(element)
            ).click()
            return True
        except Exception as e:
#            print(f"[RETRY {i+1}/{retries}] Не удалось кликнуть по элементу: {str(e)}")
            if i == retries - 1:
                raise
            time.sleep(delay)

def toggle_to_mcap(driver):
    try:
#        print("[DEBUG]  Переключатель Price → MCap...")

        # Прокручиваем вверх и ждём отрисовки вкладок
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(1)

        # Проверка: уже включён MCap?
        try:
            tabs = WebDriverWait(driver, 5).until(
                EC.presence_of_all_elements_located((By.XPATH, "//div[contains(@class, 'chakra-tabs__tab')]"))
            )
            for tab in tabs:
                selected = tab.get_attribute("aria-selected")
                text = tab.text.strip().upper()
                if selected == "true":
#                    print(f"[DEBUG] Активный режим: {text}")
                    if "MCAP" in text:
#                        print("[DEBUG] MCap уже активен — переключение не требуется")
                        return  # Выходим, не дожидаясь Price
        except Exception as e:
            print(f"[WARN] Не удалось определить текущий режим: {e}")

        print("[DEBUG] ⏳ Ожидаем блок с текстом Price...")

        price_header = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,
                "//div[contains(@class, 'text-sm') and contains(text(), 'Price')]"
            ))
        )

        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", price_header)
        time.sleep(1)

        ActionChains(driver).move_to_element(price_header).pause(0.5).click().perform()

        print("[SUCCESS] ✅ Переключено на MCap")
        time.sleep(2)

    except Exception as e:
        print(f"[WARN] ⚠️ Не удалось переключить на MCap: {e}")

def grab_toolbar_text(driver, step, toolbar_ref_x, toolbar_ref_y):
    screenshot = driver.get_screenshot_as_png()
    img = Image.open(BytesIO(screenshot))

    # Обрезка
    crop_left = toolbar_ref_x + 70
    crop_top = toolbar_ref_y + 20
    crop_right = toolbar_ref_x + 180
    crop_bottom = crop_top + 15
    crop = img.crop((crop_left, crop_top, crop_right, crop_bottom))

    # Увеличение сразу
    crop = crop.resize((crop.width * 6, crop.height * 6), Image.LANCZOS)
    crop = crop.filter(ImageFilter.GaussianBlur(radius=0.8))  # помогает разлипанию

    # Перевод в ч/б + инверсия
    crop = crop.convert("L")
    crop = ImageOps.invert(crop)

    # Контраст и резкость
    crop = ImageEnhance.Contrast(crop).enhance(1.5)
    crop = ImageEnhance.Sharpness(crop).enhance(1.0)

    # Бинаризация
    np_img = np.array(crop)
    threshold = 140
    np_img = np.where(np_img > threshold, 255, 0).astype(np.uint8)
    crop = Image.fromarray(np_img)

    # Утолщаем шрифт (осторожно)
    crop = crop.filter(ImageFilter.MaxFilter(size=1))

    # Сохраняем
    path = f"/home/chromeuser/screens/mcap_history/crop_toolbar_{step:03d}.png"
    crop.save(path)

    # OCR
    config = "--psm 6 -c tessedit_char_whitelist='0123456789abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ.,KMBHLOC'"
    text = pytesseract.image_to_string(crop, config=config)

    return text.strip()

def grab_timeline_text(driver, toolbar_ref_x, toolbar_ref_y, step):
    screenshot = driver.get_screenshot_as_png()
    img = Image.open(BytesIO(screenshot)).convert("RGB")

    # ������ Чуть выше и толще прямоугольник
    crop_top = toolbar_ref_y + 900 #если - то наверх если + то вниз двигает окно целиком ввер вниз
    crop_bottom = crop_top + 19 #ширина растояния по оси y расширяет окно
    crop_right = toolbar_ref_x + 2163 #если + то вправо
    crop_left = crop_right - 110

    block = img.crop((crop_left, crop_top, crop_right, crop_bottom))

    # Сохраняем изображение
    save_path = f"/home/chromeuser/screens/mcap_history/timeline_datebox_{step:03d}.png"
    block.save(save_path)
#    print(f"[MCAP]  Сохранили вырез даты: {save_path}")

    # OCR: усиление + увеличение
    block_gray = ImageOps.invert(block.convert("L"))
    block_gray = ImageEnhance.Contrast(block_gray).enhance(3.5)
    block_gray = ImageEnhance.Sharpness(block_gray).enhance(2.0)
    block_gray = block_gray.resize((block_gray.width * 3, block_gray.height * 3), Image.LANCZOS)

    config = "--psm 7 -c tessedit_char_whitelist='0123456789abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ‘’' "
    ocr_text = pytesseract.image_to_string(block_gray, config=config).strip()

    # ������ Логируем для анализа
    with open("/home/chromeuser/screens/mcap_history/ocr_dates_debug.txt", "a") as f:
        f.write(f"{step:03d}: {ocr_text}\n")

#    print(f"[MCAP][OCR Timeline] {ocr_text}")
    return ocr_text

def activate_crosshair_on_graph(driver):
#    print("[MCAP] Ищем активный canvas...")
    canvases = driver.execute_script("""
        return Array.from(document.querySelectorAll('canvas')).map((c, i) => {
            const r = c.getBoundingClientRect();
            return {index: i, width: r.width, height: r.height, x: r.left, y: r.top, display: getComputedStyle(c).display, opacity: getComputedStyle(c).opacity};
        });
    """)
    canvases = [c for c in canvases if c["width"] > 300 and c["height"] > 100 and c["display"] != "none" and float(c["opacity"]) > 0.1]
    if not canvases:
#        print("[MCAP][ERROR] ❌ Нет подходящих canvas")
        return False

    canvas_info = canvases[0]
    canvas_element = driver.find_elements(By.TAG_NAME, "canvas")[canvas_info['index']]

    safe_offset_x = int(canvas_info["width"] * 0.995)
    safe_offset_y = int(canvas_info["height"] * 0.5)

    ActionChains(driver).move_to_element_with_offset(canvas_element, safe_offset_x - canvas_info["width"]//2, safe_offset_y - canvas_info["height"]//2).pause(0.5).click().perform()
    time.sleep(1)
#    print(f"[MCAP] ✅ Клик в очень правую часть canvas ({safe_offset_x}px, {safe_offset_y}px) — перекрестие установлено")
    return True

def parse_high_from_toolbar(toolbar_text: str, previous_value: str | None = None) -> str | None:
    if not toolbar_text:
#        print("[H-EXTRACT] Пустой входной текст")
        return None

    # Фиксируем определённые символы
    fix_map = {
        "O": "0", "o": "0", "l": "1", ",": ".", "‘": "", "’": "", "“": "", "”": ""
    }
    text_fixed = ''.join(fix_map.get(c, c) for c in toolbar_text)

    # Ищем сегмент H с возможными суффиксами K, M, B
    match = re.search(r"[A-Z]{0,3}?(H[\dA-Z\.]{2,10}[KMB])", text_fixed)
    if not match:
#        print("[H-EXTRACT] Не найден H-сегмент в:", text_fixed)
        return None

    raw_h = match.group(1)

    # Обрезаем текст от "H" до суффикса (K, M, B)
    m = re.match(r"H([^\.\s]+)\.(\w{2,4})([KMB])", raw_h)
    if not m:
#        print("[H-EXTRACT] Неполный формат:", raw_h)
        return None

    int_part_raw, frac_part, suffix_raw = m.groups()

    suffix = {"K": "K", "M": "M", "B": "B"}.get(suffix_raw, "")
    use_replacement = False

    # Функция для преобразования строки в число
    def try_convert(s):
        try:
            return float(s)
        except:
            return None

    # Функция для получения множителя суффикса
    def suffix_multiplier(sfx):
        return {"K": 1e3, "M": 1e6, "B": 1e9}.get(sfx, 1)

    # Логика замены S в дробной части
    if any(c in frac_part for c in "SsIi"):
        replaced = frac_part.replace("S", "5").replace("s", "5").replace("I", "1").replace("i", "1")
        if re.fullmatch(r"\d{2}", replaced):
#            print(f"[H-FIX] Заменили S в дробной части: {frac_part} → {replaced}")
            frac_part = replaced
        else:
            digits_only = ''.join(c for c in replaced if c.isdigit())
            if len(digits_only) >= 2:
#                print(f"[H-FIX] Удаляем лишние символы и обрезаем дробную часть: {frac_part} → {digits_only[:2]}")
                frac_part = digits_only[:2]
            else:
#                print(f"[H-EXTRACT] ❌ Невозможно корректно исправить дробную часть: {frac_part}")
                return None

    # Обрезаем дробную часть до двух цифр
    if len(frac_part) > 2:
        digits = ''.join(c for c in frac_part if c.isdigit())
        if len(digits) >= 2:
#            print(f"[H-FIX] Обрезаем дробную часть до двух цифр: {frac_part} → {digits[:2]}")
            frac_part = digits[:2]
        else:
            return None

    # Переносим одну цифру из целой части в дробную, если необходимо
    int_digits = re.sub(r"[^\d]", "", int_part_raw)
    if len(int_digits) == 4 and len(frac_part) == 1:
#        print(f"[H-FIX] Переносим цифру из int в frac: {int_digits}.{frac_part} → {int_digits[:-1]}.{int_digits[-1]}{frac_part}")
        frac_part = int_digits[-1] + frac_part
        int_part_raw = int_digits[:-1]

    # Удаляем ведущие нули в целой части
    if int_part_raw.startswith("00"):
#        print(f"[H-FIX] Удаляем один из двух ведущих нулей: {int_part_raw} → {int_part_raw[1:]}")
        int_part_raw = int_part_raw[1:]

    # Проверяем и заменяем S в целой части
    if any(c in int_part_raw for c in "SsIi"):
        digits_only = re.sub(r"[^\d]", "", int_part_raw)
        temp = int_part_raw
        if digits_only.startswith("00"):
            first_zero_index = temp.find("0")
            if first_zero_index != -1:
                temp = temp[:first_zero_index] + temp[first_zero_index+1:]
#                print(f"[H-FIX] Удалён 0 до замены S: {int_part_raw} → {temp}")
        test_with_5 = temp.replace("S", "5").replace("s", "5").replace("I", "1").replace("i", "1")
        test_clean = re.sub(r"[^\d]", "", test_with_5)
        test_value_str = f"{test_clean}.{frac_part}{suffix}"
        current_full = try_convert(f"{test_clean}.{frac_part}") * suffix_multiplier(suffix)
        if previous_value and re.match(r"H\d+\.\d{2}[KMB]", previous_value):
            prev_number = float(previous_value[1:-1])
            prev_suffix = previous_value[-1]
            prev_full = prev_number * suffix_multiplier(prev_suffix)
            ratio = current_full / prev_full if prev_full > 0 else 1
            if 0.2 <= ratio <= 5:
                use_replacement = True
                int_part_clean = test_clean
#                print(f"[H-FIX] S→5 заменено: {int_part_raw}.{frac_part}{suffix} → {int_part_clean}.{frac_part}{suffix} (разница: {ratio:.2f}x)")
            else:
                pass  # или твой код
        else:
            pass  # или твой код

    if not use_replacement:
        int_part_clean = re.sub(r"[^\d]", "", int_part_raw)

    # Удаляем лишний ведущий 0
    if len(int_part_clean) > 3 and int_part_clean.startswith("0"):
#        print(f"[H-FIX] Удаляем лишний ведущий 0: {int_part_clean} → {int_part_clean[1:]}")
        int_part_clean = int_part_clean[1:]

    if len(int_part_clean) > 3:
#        print(f"[H-EXTRACT] Целая часть слишком длинная: {int_part_clean}")
        return None

    # Финальный результат
    final = f"H{int_part_clean}.{frac_part}{suffix}"
    return final

def clean_raw_h_segment(segment_text):
   """
   Очищает сырой сегмент между H и L, возвращает строку вида 'H7.25K'.
   """
   if not segment_text:
       return None

   fix_map = {"S": "5", "s": "5", "O": "0", "o": "0", "I": "1", "l": "1"}
   cleaned = ''.join(fix_map.get(c, c) for c in segment_text)
   cleaned = re.sub(r"[^\d\.KMB]", "", cleaned)  # оставляем только цифры, точку и суффиксы

   if not re.search(r"\d", cleaned):
       return None

   return "H" + cleaned

def normalize_date_text(text: str) -> str:
   if not text:
       return ""
       
   # Жёсткая замена: если день выглядит как S0 перед месяцем — это 30
   text = re.sub(r"\bS0\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\b", r"30 \1", text, flags=re.IGNORECASE)
   text = re.sub(r"\bF\b\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\b", r"5 \1", text, flags=re.IGNORECASE)

   # Общие OCR-фиксы
   replacements_general = {
       "‘": "'", "’": "'", "“": "'", "”": "'", "?": "'", '"': "'", "•": "",
       "у": "y",  # кириллическая y
   }
   replacements_digits = {
       "O": "0", "o": "0", "Q": "0", "D": "0",
       "I": "1", "l": "1", "|": "1", "!": "1",
       "Z": "2", "z": "2",
       "B": "8", "s": "5"
   }
   for wrong, right in replacements_general.items():
       text = text.replace(wrong, right)
   for wrong, right in replacements_digits.items():
       text = text.replace(wrong, right)

   # Артефакты OCR в названиях месяцев
   MONTH_OCR_FIXES = {
       # Январь
       "jal": "jan", "jav": "jan",
       # Февраль
       "fev": "feb", "fe6": "feb", "fe5": "feb", "fep": "feb",
       # Март
       "m4r": "mar", "mae": "mar",
       # Апрель
      "ap": "apr", "an": "apr", "am": "apr", "ar": "apr",
      "amp": "apr", "apl": "apr", "aprll": "apr", "aр": "apr", "aл": "apr",
       # Май
       "maу": "may", "mav": "may", "mayy": "may",
       # Июнь
       "ju6": "jun", "j6n": "jun",
       # Июль
       "ju1": "jul", "jui": "jul", "jly": "jul",
       # Август
       "auq": "aug", "4ug": "aug", "aqg": "aug",
       # Сентябрь
       "5ep": "sep", "scp": "sep",
       # Октябрь
       "0ct": "oct", "octt": "oct",
       # Ноябрь
       "n0v": "nov", "noy": "nov",
       # Декабрь
       "d3c": "dec", "dee": "dec",
   }

   for wrong, correct in MONTH_OCR_FIXES.items():
       text = re.sub(fr"\b{wrong}\b", correct, text, flags=re.IGNORECASE)

   # Частные случаи OCR
   text = re.sub(r"\bOF([A-Za-z]{3,})", r"01 \1", text, flags=re.IGNORECASE)
   text = re.sub(r"\bO1([A-Za-z]{3,})", r"01 \1", text, flags=re.IGNORECASE)
   text = re.sub(r"\bO(\d)([A-Za-z]{3,})", r"0\1 \2", text, flags=re.IGNORECASE)
   text = re.sub(r"\bju1\b", "jul", text, flags=re.IGNORECASE)
   text = re.sub(r"\bA[mpр]+[rp]?\b", "Apr", text, flags=re.IGNORECASE)

   # Удаляем всё лишнее
   text = re.sub(r"[^\w\s]", "", text)
   text = re.sub(r"\s+", " ", text).strip()

   # Разделяем дату и слипшееся время
   text = re.sub(r"(\d{2})(\d{4})$", r"\1 \2", text)
   text = re.sub(r"(\d{3})(\d{4})$", r"\1 \2", text)
   text = re.sub(r"(\d{1,2})([A-Za-z]{3,})(\d{2})", r"\1 \2 \3", text)
   text = re.sub(r"\b0(\d)\b", r"\1", text)

   return text.strip()

def normalize_time_part(hour: str, minute: str) -> (str, str):
    replacements = {
        "O": "0", "o": "0", "Q": "0", "D": "0",
        "I": "1", "l": "1", "|": "1", "!": "1",
        "S": "5", "s": "5"
    }

    hour = ''.join(replacements.get(c, c) for c in hour)
    minute = ''.join(replacements.get(c, c) for c in minute)

    # Только цифры
    hour = re.sub(r"\D", "", hour)
    minute = re.sub(r"\D", "", minute)

    hour = hour.zfill(2)
    minute = minute.zfill(2)

    # Ограничения диапазона
    if int(hour) > 23:
        hour = "23"
    if int(minute) > 59:
        minute = "59"

    return hour, minute

def clean_and_parse_date(raw_text):
    if not raw_text:
        return None

    cleaned = normalize_date_text(raw_text)
#    print(f"[MCAP][DATE raw→clean] {cleaned}")

    MONTH_FIXES = {
    # Январь
    "jan": 1, "jav": 1, "ja": 1, "jal": 1,
    # Февраль
    "feb": 2, "fev": 2, "fe": 2, "fep": 2, "fe6": 2, "fe5": 2,
    # Март
    "mar": 3, "m4r": 3, "mae": 3,
    # Апрель
    "apr": 4, "aprll": 4, "apl": 4, "arp": 4, "amp": 4,
    # Май
    "may": 5, "maу": 5, "mav": 5, "mayy": 5,
    # Июнь
    "jun": 6, "ju6": 6, "j6n": 6,
    # Июль
    "jul": 7, "ju1": 7, "jui": 7, "jly": 7,
    # Август
    "aug": 8, "auq": 8, "4ug": 8, "aqg": 8,
    # Сентябрь
    "sep": 9, "5ep": 9, "scp": 9,
    # Октябрь
    "oct": 10, "0ct": 10, "octt": 10,
    # Ноябрь
    "nov": 11, "n0v": 11, "noy": 11,
    # Декабрь
    "dec": 12, "d3c": 12, "dee": 12,
    }

    # Пробуем: день, месяц, 3-значный "год" и 4 цифры (на самом деле 2+время)
    match3 = re.search(r"(\d{1,2})\s*([A-Za-z]{3,})\s*(\d{3})\s*(\d{3,4})", cleaned)
    if match3:
        day, month_abbr, year3, time_extra = match3.groups()
        year_clean = "20" + year3[:2]
        time_part_raw = year3[2] + time_extra
    else:
        # Обычный случай: день, месяц, 2-значный год + 4–5 цифр времени
        match2 = re.search(r"(\d{1,2})\s*([A-Za-z]{3,})\s*(\d{2,4})\s*(\d{4,5})$", cleaned)
        if match2:
            day, month_abbr, year2, time_part_raw = match2.groups()
    
            # Если год ошибочно содержит 4 цифры (например, 2591)
            if len(year2) == 4 and year2.startswith("25") and "9" in year2[2:]:
                rest = year2[2:].replace("9", "")
                if len(rest) == 1:
#                    print(f"[MCAP][FIX] Удалили артефакт 9 из года и перенесли цифру во время: год={year2} → 25, переносим {rest} в начало времени")
                    year_clean = "2025"
                    time_part_raw = rest + time_part_raw
                else:
                    year_clean = "20" + year2[:2]  # просто 20 + первые 2 цифры
            else:
                year_clean = "20" + year2
        else:
#            print(f"[MCAP][ERROR] Не удалось разделить дату и время: {cleaned}")
            return None

    # Обработка артефакта '7' вместо ':' — если время 5 цифр и 3-я = '7'
    if len(time_part_raw) == 5 and time_part_raw[2] == '7':
        hour = time_part_raw[:2]
        minute = time_part_raw[3:]
    elif len(time_part_raw) >= 4:
        hour = time_part_raw[:2]
        minute = time_part_raw[2:4]
    else:
        hour, minute = "00", "00"
    
    hour, minute = normalize_time_part(hour, minute)
    
    # Очистка дня
    if len(day) > 2 and day.startswith("0"):
        day = day[1:]
    
    # ������ Новый блок: если день > 31 — попробуем урезать
    if int(day) > 31:
        if len(day) == 3 and day.startswith("1"):
#            print(f"[MCAP][FIX] Урезали некорректный день: {day} → {day[:2]}")
            day = day[:2]
        elif len(day) == 3 and day[0] == '3':
#            print(f"[MCAP][FIX] Приняли максимум: {day} → 31")
            day = "31"
        elif len(day) == 3 and day[0] == '0':
#            print(f"[MCAP][FIX] Убрали первый 0: {day} → {day[1:]}")
            day = day[1:]
        else:
#            print(f"[MCAP][WARN] ❌ Некорректный день: {day}")
            return None
    
    if int(day) == 0:
        day = "30"  # OCR мог распознать 30 как "S0" → "00" после замен

    month_abbr_clean = re.sub(r"[^A-Za-z]", "", month_abbr).lower()[:3]
    month = MONTH_FIXES.get(month_abbr_clean)
    
    if not month:
        import difflib
        close = difflib.get_close_matches(month_abbr_clean, list(MONTH_FIXES.keys()), n=1, cutoff=0.6)
        if close:
            month = MONTH_FIXES[close[0]]

    if not month:
#        print(f"[MCAP][WARN] Не удалось определить месяц: {month_abbr_clean}")
        return None

    try:
        return datetime.strptime(f"{year_clean}-{month:02d}-{int(day):02d} {hour}:{minute}", "%Y-%m-%d %H:%M")
    except Exception as e:
#        print(f"[MCAP][ERROR] Ошибка парсинга даты: {e}")
        return None

def test_mcap_scraper(driver, toolbar_ref_x, toolbar_ref_y, wallet_index=0, token_index=0, stop_at_date=None):
    save_dir = "/home/chromeuser/screens/mcap_history"
    os.makedirs(save_dir, exist_ok=True)
    history = []
    log_entries = []

#    print("[MCAP] Сброс графика (Alt+R)...")
    ActionChains(driver).key_down(Keys.ALT).send_keys('r').key_up(Keys.ALT).perform()
    time.sleep(2)

#    print("[MCAP] Активируем перекрестие на графике...")
    if not activate_crosshair_on_graph(driver):
        return None

    step = 0
    previous_h_text = None  # для передачи в parse_high_from_toolbar
    empty_candle_count = 0
    time_window = deque(maxlen=5)

    while True:
        toolbar_text = grab_toolbar_text(driver, step, toolbar_ref_x, toolbar_ref_y)
        timeline_date = grab_timeline_text(driver, toolbar_ref_x, toolbar_ref_y, step)

#        print(f"[MCAP][TOOLBAR OCR] {toolbar_text}")
#        print(f"[MCAP][TIMELINE OCR raw] {timeline_date}")

#        try:
#            os.remove(f"/home/chromeuser/screens/mcap_history/crop_toolbar_{step:03d}.png")
#            os.remove(f"/home/chromeuser/screens/mcap_history/timeline_datebox_{step:03d}.png")
#        except Exception as e:
#            print(f"[WARN] Не удалось удалить скриншоты на шаге {step:03d}: {e}")

        parsed_date = None
        timeline_text_clean = normalize_date_text(timeline_date)
#        print(f"[MCAP][TIMELINE OCR clean] {timeline_text_clean}")

        # Спец-кейсы OCR
        m = re.fullmatch(r"(?:of|0f)?jul(\d{2})", timeline_text_clean.lower())
        if m:
            parsed_date = datetime.strptime(f"20{m.group(1)}-07-07", "%Y-%m-%d")
        else:
            m = re.fullmatch(r"31jul(\d{2})", timeline_text_clean.lower())
            if m:
                parsed_date = datetime.strptime(f"20{m.group(1)}-07-31", "%Y-%m-%d")
            else:
                try:
                    parsed_date = clean_and_parse_date(timeline_text_clean)
                except Exception as e:
                    pass  # или твой код

        # ������ Проверка: достигли даты из кэша?
        if stop_at_date and parsed_date and parsed_date <= stop_at_date:
#            print(f"[MCAP] ✅ Достигли даты кэша: {parsed_date} ≤ {stop_at_date} — остановка прокрутки")
            break

        time_window.append(timeline_date)
        if len(time_window) == 5 and len(set(time_window)) == 1:
#            print(f"[MCAP] ❌ График завершён. Найдены 5 одинаковых временных меток: {time_window[0]}")
            break

#        is_toolbar_empty = bool(re.search(r"H\s*([\-0O〇]|$)", toolbar_text.strip()))
#        if is_toolbar_empty and not parsed_date:
#            empty_candle_count += 1
#            print(f"[MCAP][CHECK] Пустая свеча: {empty_candle_count}/3")
#        else:
#            empty_candle_count = 0
#
#        if empty_candle_count >= 3:
#            print("[MCAP] Обнаружено 3 подряд пустые свечи — завершаем парсинг.")
#            break

        high_value_text = parse_high_from_toolbar(toolbar_text, previous_value=previous_h_text)
        date_for_log = parsed_date.strftime("%Y-%m-%d") if parsed_date else (timeline_date if timeline_date else "unknown")

        if parsed_date and high_value_text:
            history.append((parsed_date, high_value_text))
            previous_h_text = high_value_text  # <== ДОБАВЛЕНО
            log_path = os.path.join(save_dir, "mcap_history_log.txt")
#            print(f"[MCAP][H-VALUE ✅] Чистое H-значение для лога: {high_value_text}")
            with open(log_path, "a") as f:
                f.write(f"{parsed_date.strftime('%Y-%m-%d %H:%M')} — {high_value_text}\n")
#            print(f"[MCAP][LOG] {date_for_log} — {high_value_text}")
        else:
            pass  # или твой код

        ActionChains(driver).send_keys(Keys.ARROW_LEFT).perform()
        time.sleep(random.uniform(0.01, 0.011))
        step += 1
        if step > 50000:
#            print("[MCAP] ❌ Превышено максимальное количество шагов")
            break

    mcap_timeline = []
    for item in history:
        if not isinstance(item, tuple) or len(item) != 2:
#            print(f"[MCAP][FIX] ❌ Некорректный элемент в history: {item}")
            continue
        date_obj, val = item
        if not isinstance(date_obj, datetime) or not isinstance(val, str):
#            print(f"[MCAP][FIX] ❌ Неверные типы: date={type(date_obj)}, val={type(val)}")
            continue
        try:
            h_val = parse_high_from_toolbar(val)
            if h_val:
                mcap_timeline.append((date_obj, h_val))
        except Exception as e:
#            print(f"[MCAP][ERROR]  Ошибка разбора toolbar: {e}")
            continue
    
    return mcap_timeline

def run_detailed_mcap_scraper(driver, contract_address=None):
    try:
        stop_at_date = None
        existing_timeline = []

        # === 1. Проверка кэша ===
        if contract_address:
            filepath = f"/home/chromeuser/screens/token_cache/{contract_address}.json"
            existing_timeline = []
            if os.path.exists(filepath):
                with open(filepath, "r") as f:
                    raw_data = json.load(f)
                existing_timeline = [(datetime.strptime(k, "%Y-%m-%d %H:%M"), v) for k, v in raw_data.items()]
            if existing_timeline:
                latest_cached_date = max(dt for dt, _ in existing_timeline)
                now = datetime.utcnow()

                if now - latest_cached_date < timedelta(hours=1):
#                    print(f"[CACHE] ✅ Актуальный кэш найден для {contract_address} — пропускаем анализ")
                    return existing_timeline
                else:
                    stop_at_date = latest_cached_date
#                    print(f"[CACHE] ⚠ Найден неполный кэш — дополняем от {latest_cached_date} до {now}")

        # === 2. Выходим из iframe, получаем координаты графика ===
        driver.switch_to.default_content()
        driver.execute_script("window.scrollTo(0, 0);")
        driver.execute_script("window.scrollBy(0, 200);")
        time.sleep(0.3)
        driver.execute_script("window.scrollBy(0, -200);")
        driver.execute_script("window.dispatchEvent(new Event('resize'));")
        time.sleep(0.7)

#        print("[MCAP] Ждём появления якоря H1...")
        timeframe_anchor = WebDriverWait(driver, 15).until(
            EC.visibility_of_element_located((By.XPATH, "//span[text()='1H' and contains(@class, 'cursor-pointer')]"))
        )
        time.sleep(2.0)
        # Кликаем по таймфрейму 1h
        ActionChains(driver).move_to_element(timeframe_anchor).pause(0.3).click().perform()
#        print("[MCAP] ✅ Переключились на таймфрейм 1h")
        time.sleep(2.0)
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", timeframe_anchor)
        time.sleep(1.0)
        rect = driver.execute_script("""
            const el = arguments[0];
            const rect = el.getBoundingClientRect();
            return { x: rect.left, y: rect.bottom };
        """, timeframe_anchor)
        toolbar_ref_x, toolbar_ref_y = int(rect["x"]), int(rect["y"])
#        print(f"[MCAP] Координаты toolbar: ({toolbar_ref_x}, {toolbar_ref_y})")

        # === 3. Возвращаемся в iframe для анализа графика ===
        iframe = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.TAG_NAME, "iframe")))
        driver.switch_to.frame(iframe)
#        print("[MCAP] ✅ Вернулись в iframe для анализа графика")

        # === 4. Запуск анализа графика ===
#        print("[MCAP]  Запуск test_mcap_scraper...")
        mcap_timeline_new = test_mcap_scraper(
            driver,
            toolbar_ref_x,
            toolbar_ref_y,
            stop_at_date=stop_at_date
        )

        if mcap_timeline_new is None:
#            print(f"[ERROR] ❌ Ошибка при анализе графика — пропускаем токен")
            return []

        # === 4.1 Проверка формата результата
        if not isinstance(mcap_timeline_new, list) or not all(
            isinstance(x, tuple) and len(x) == 2 and isinstance(x[0], datetime) and isinstance(x[1], str)
            for x in mcap_timeline_new
        ):
#            print(f"[ERROR] ❌ Невалидный формат mcap_timeline_new — пропускаем токен")
            return []

        # === 4.2 Проверка — есть ли вообще новые даты
        if stop_at_date and all(dt <= stop_at_date for dt, _ in mcap_timeline_new):
#            print(f"[CACHE] Нет новых данных после {stop_at_date} — дописывать нечего")
            return existing_timeline

        # === 5. Объединяем с кэшем и сохраняем ===
        if contract_address:
            combined = {dt.strftime("%Y-%m-%d %H:%M"): val for dt, val in existing_timeline}
            for dt, val in mcap_timeline_new:
                key = dt.strftime("%Y-%m-%d %H:%M")
                if key not in combined:
                    combined[key] = val
            combined_timeline = sorted([
                (datetime.strptime(k, "%Y-%m-%d %H:%M"), v)
                for k, v in combined.items()
            ])
            filepath = f"/home/chromeuser/screens/token_cache/{contract_address}.json"
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            with open(filepath, "w") as f:
                json.dump({
                    dt.strftime("%Y-%m-%d %H:%M"): val
                    for dt, val in combined_timeline
                }, f, indent=2)
#            print(f"[HISTORY] ✅ История токена сохранена: {filepath}")
            return combined_timeline

        return mcap_timeline_new

    except Exception as e:
#        print(f"[MCAP] ❌ Ошибка в run_detailed_mcap_scraper: {e}")
        return []

    finally:
        driver.switch_to.default_content()
        time.sleep(1)

def navigate_to_wallet_via_search(driver, wallet):
    try:
#        print(f"[NAVIGATE] ▶ Открытие поиска и ввод кошелька: {wallet}")
        
        # Прокручиваем вверх, на всякий случай
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(0.5)

        # Сначала пробуем найти по placeholder, затем fallback
        try:
            search_input = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Search token/contract/wallet']"))
            )
        except:
            search_input = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='text']"))
            )

        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", search_input)
        ActionChains(driver).move_to_element(search_input).pause(0.4).click().perform()
        time.sleep(0.5)

        # Удаляем старое значение (через JS и .clear)
        driver.execute_script("arguments[0].value = '';", search_input)
        search_input.clear()
        time.sleep(0.3)

        search_input.send_keys(wallet)
        time.sleep(1.0)

        # Повторная проверка ввода
        if search_input.get_attribute("value") != wallet:
#            print("[WARNING] send_keys не сработал — вставляем через JS")
            driver.execute_script("arguments[0].value = arguments[1];", search_input, wallet)
            time.sleep(1.0)

#        print("[NAVIGATE] ⏳ Ждем появления результата...")

        result_xpath = f"//a[contains(@href, '/sol/address/{wallet}')]"
        result_link = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, result_xpath))
        )

        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", result_link)
        time.sleep(random.uniform(0.6, 1.0))

        # Принудительный "сброс мышки", чтобы избежать бага move target out of bounds
        driver.execute_script("window.scrollBy(0, -100);")
        ActionChains(driver).move_by_offset(1, 1).perform()
        time.sleep(0.3)

        # Проверка перекрытия
        overlapping = driver.execute_script("""
            const rect = arguments[0].getBoundingClientRect();
            const x = rect.left + rect.width / 2;
            const y = rect.top + rect.height / 2;
            const el = document.elementFromPoint(x, y);
            return el !== arguments[0] ? el.outerHTML : null;
        """, result_link)
        if overlapping:
            pass  # или твой код

        # Клик по ссылке
        try:
            ActionChains(driver).move_to_element(result_link).pause(0.4).click().perform()
            print("[NAVIGATE] ✅ Клик по кошельку выполнен через ActionChains")
        except Exception as e:
            print(f"[WARNING] Ошибка ActionChains: {e} — пробуем JS")
            driver.execute_script("arguments[0].click();", result_link)

        print("[NAVIGATE] ⏳ Ждем загрузки страницы кошелька...")
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, "//div[contains(text(), 'PnL') or contains(text(), 'Bal')]"))
        )
        time.sleep(random.uniform(1.0, 1.6))
        print("[NAVIGATE] ✅ Страница кошелька загружена")

    except Exception as e:
        print(f"[NAVIGATE][CRITICAL] ❌ Ошибка перехода к кошельку: {e}")

def process_wallet(driver, wallet, wallet_index=0, token_resume_index=0, resume_stage="start"):
    results = []
    try:
        print(f"[INFO] Обработка кошелька: {wallet}")
        update_resume_state(wallet_index=wallet_index, token_index=token_resume_index, stage="start")
        # Переход напрямую по адресу кошелька
        navigate_to_wallet_via_search(driver, wallet)
        solve_turnstile_if_present(driver)
        # Закрываем возможный модал "I Know"
        close_pnl_modal_if_present(driver)
        time.sleep(random.uniform(2.5, 4.5))  # Человеческая задержка

        # Проверяем баланс
        try:
            print("[INFO] Проверка баланса кошелька...")
            bal_elem = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//div[contains(text(), 'Bal')]/following-sibling::div"))
            )
            bal_text = bal_elem.text.strip().split("\n")[0].replace("SOL", "").strip()
            balance_sol = float(bal_text)
            if balance_sol == 0:
                print("[INFO] Баланс = 0 SOL — пропускаем кошелек")
                return results
            print(f"[INFO] Баланс кошелька: {balance_sol:.2f} SOL")
            update_resume_state(wallet_index=wallet_index, token_index=token_resume_index, stage="balance_checked")
        except Exception as e:
            print(f"[ERROR] Не удалось определить баланс: {e}")
            return results

        # Извлекаем до 30 уникальных названий токенов из Recent PnL
        try:
            print("[INFO] Поиск токенов в Recent PnL...")
        
            # Кликаем по якорю "All"
            try:
                all_anchor = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//div[text()='All' and contains(@class, 'cursor-pointer')]"))
                )
                random_mouse_move(driver)
                random_scroll(driver)
                ActionChains(driver).move_to_element(all_anchor).pause(random.uniform(0.3, 0.6)).perform()
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", all_anchor)
                time.sleep(random.uniform(1.0, 2.0))
                all_anchor.click()
                time.sleep(random.uniform(0.5, 1.0))
                print("[INFO] ✅ Клик по якорю 'All' выполнен")
            except Exception as e:
                print(f"[WARN] ⚠️ Якорь 'All' не найден или не кликабелен: {e}")
        
            update_resume_state(wallet_index=wallet_index, token_index=token_resume_index, stage="clicked_all")
        
            # Фокус на теле страницы перед прокруткой
            try:
                body = driver.find_element(By.TAG_NAME, "body")
                ActionChains(driver).move_to_element(body).pause(0.2).perform()
            except Exception as e:
                print(f"[WARN] ⚠️ Не удалось навести на body: {e}")
        
            # ������ Физическая прокрутка (эмулирует колесо мыши)
            pyautogui.scroll(-600)  # вниз
            time.sleep(random.uniform(0.7, 1.3))
            pyautogui.scroll(400)   # вверх
            time.sleep(random.uniform(0.4, 0.7))
        
            # Ожидаем таблицу
            recent_table = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//div[@id='tabs-leftTabs--tabpanel-0']//table"))
            )
            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", recent_table)
            time.sleep(2)
        
            # Сохраняем token_name + link
            token_links = recent_table.find_elements(By.XPATH, ".//tbody//tr//td[1]//a")
            token_data_list = []
            for link in token_links:
                name = link.text.strip().split("\n")[0]
                href = link.get_attribute("href")
                if name and name not in [t['name'] for t in token_data_list]:
                    token_data_list.append({'name': name, 'link': href})
                if len(token_data_list) >= 15:
                    break
        
            if not token_data_list:
                print("[INFO] ⚠️ Токены не найдены в Recent PnL — пропуск")
                return results
        
            print(f"[INFO] ✅ Найдено {len(token_data_list)} уникальных токенов: {[t['name'] for t in token_data_list]}")
            update_resume_state(wallet_index=wallet_index, token_index=token_resume_index, stage="token_table_loaded")
        
            print(f"[INFO]  Начинаем перебор {len(token_data_list)} токенов...")
        
            for i in range(token_resume_index, len(token_data_list)):
                token = token_data_list[i]
                token_name = token['name']
                token_link = token['link']

                try:
                    update_resume_state(wallet_index=wallet_index, token_index=i, stage="token_start")
                    print(f"[INFO] ▶ Обрабатываем токен {i+1}: {token_name}")
        
                    # Удаляем старые скриншоты
                    for f in glob.glob("/home/chromeuser/screens/crop_toolbar_*.png"):
                        os.remove(f)
                    for f in glob.glob("/home/chromeuser/screens/crop_toolbar_left_*.png"):
                        os.remove(f)
        
                    random_mouse_move(driver)
                    random_scroll(driver)
        
                    # Переход по сохранённой ссылке
                    driver.get(token_link)
                    time.sleep(random.uniform(2.0, 3.0))
        
                    solve_turnstile_if_present(driver)
                    update_resume_state(wallet_index=wallet_index, token_index=i, stage="token_clicked")
        
                except Exception as e:
                    print(f"[ERROR] ⚠️ Ошибка обработки строки {i+1}: {e}")
                    continue
                
                print("[SCROLL] Попытка прокрутки к фильтру Maker через JS")

                try:
                    maker_filter_elem = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, "//*[text()='Maker']"))
                    )
                    driver.execute_script("arguments[0].scrollIntoView({behavior: 'instant', block: 'center'});", maker_filter_elem)
                    ActionChains(driver).move_to_element(maker_filter_elem).pause(0.5).perform()
                    print("[SCROLL] ✅ Успешно прокрутили к фильтру Maker")
                    time.sleep(1.5)
                except Exception as e:
                    print(f"[WARNING] ❌ Не удалось найти или прокрутить к фильтру Maker: {e}")


                # Клик по нужной иконке-фильтра справа от Maker
                try:
                    print("[SCROLL] Попытка прокрутки к фильтру Maker через XPath")
                
                    # Найдём сам блок с Maker
                    maker_block = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'css-lox6ag') and contains(., 'Maker')]"))
                    )
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", maker_block)
                    time.sleep(0.3)
                    
                    # Микро-прокрутка
                    driver.execute_script("window.scrollBy(0, 30);")
                    time.sleep(0.2)
                    driver.execute_script("window.scrollBy(0, -30);")
                    time.sleep(0.2)
                    
                    print("[DEBUG] ✅ Найден блок Maker — ищем кнопку внутри")
                    
                    # ������ Теперь ищем кнопку, а не svg
                    filter_button = WebDriverWait(maker_block, 5).until(
                        EC.element_to_be_clickable((By.XPATH, ".//button[contains(@class, 'chakra-button')]"))
                    )
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", filter_button)
                    time.sleep(0.3)
                    
                    # Проверка перекрытия
                    overlapping = driver.execute_script("""
                        const rect = arguments[0].getBoundingClientRect();
                        const x = rect.left + rect.width / 2;
                        const y = rect.top + rect.height / 2;
                        const el = document.elementFromPoint(x, y);
                        return el !== arguments[0] ? el.outerHTML : null;
                    """, filter_button)
                    
                    if overlapping and "svg" not in overlapping:
                        print(f"[WARNING] Кнопка может быть перекрыта внешним элементом: {overlapping}")
                    else:
                        print("[DEBUG] ✅ Кнопка не перекрыта (или перекрыта дочерним элементом) — выполняем клик")
                
                    # Клик через JS
                    driver.execute_script("arguments[0].click();", filter_button)
                    print("[SUCCESS] Клик по кнопке фильтра Maker выполнен")
                    time.sleep(DELAY)
        
                except Exception as e:
                    print(f"[ERROR] ❌ Не удалось кликнуть по иконке фильтра Maker: {e}")
                    driver.save_screenshot("error_click_maker_filter.png")
                    navigate_to_wallet_via_search(driver, wallet)
                    solve_turnstile_if_present(driver)
                    close_pnl_modal_if_present(driver)
                    time.sleep(DELAY)
                    continue

                # Вводим адрес в поповер
                try:
                    popover = WebDriverWait(driver, 15).until(
                        EC.visibility_of_element_located((By.CSS_SELECTOR, "div[role='dialog'], div[data-popper-placement]"))
                    )
                    wallet_input = WebDriverWait(popover, 10).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "input.chakra-input, input[type='text']"))
                    )
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});arguments[0].style.border='2px solid #0f0';", wallet_input)
                    wallet_input.click()
                    time.sleep(0.3)
                    wallet_input.clear()
                    time.sleep(0.3)
                    wallet_input.send_keys(wallet)
                    if wallet_input.get_attribute('value') != wallet:
                        raise Exception("Текст не введён корректно")
                    time.sleep(1)
                except Exception as e:
                    print(f"[ERROR] Не удалось ввести кошелёк: {e}")
                    driver.save_screenshot("error_wallet_input.png")
                    driver.back()
                    solve_turnstile_if_present(driver)
                    time.sleep(DELAY * 5)
                    continue

                # ������ Повторный клик по input — активируем Apply
                try:
                    random_mouse_move(driver)
                    random_scroll(driver)
                    ActionChains(driver).move_to_element(wallet_input).pause(0.3).click().perform()
                    print("[DEBUG] Повторный клик по полю ввода кошелька выполнен")
                    time.sleep(0.6)
                except Exception as e:
                    print(f"[WARNING] Не удалось кликнуть по полю ввода повторно: {e}")

                # Нажимаем Apply с полным логированием
                try:
                    print("[DEBUG] Ищем кнопку Apply...")
                    apply_elements = driver.find_elements(By.XPATH, "//div[text()='Apply']")
                    print(f"[DEBUG] Найдено элементов Apply: {len(apply_elements)}")
                
                    for idx, el in enumerate(apply_elements):
                        print(f"  ├─ Apply[{idx}]: visible={el.is_displayed()}, enabled={el.is_enabled()}, text='{el.text}'")
                
                    apply_btn = None
                    for el in apply_elements:
                        if el.is_displayed() and el.is_enabled():
                            apply_btn = el
                            print("[DEBUG] ✅ Выбран отображаемый и активный элемент Apply")
                            break
                
                    if not apply_btn:
                        print("[DEBUG] ❌ Нет видимого и активного Apply — пробуем реанимировать попап")
                
                        # Двигаем экран, чтобы оживить попап
                        driver.execute_script("window.scrollBy(0, 200);")
                        time.sleep(0.3)
                        driver.execute_script("window.scrollBy(0, -200);")
                        time.sleep(0.5)
                
                        # Наводим на Reset
                        reset_btn = WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.XPATH, "//div[text()='Reset']"))
                        )
                        random_mouse_move(driver)
                        random_scroll(driver)
                        ActionChains(driver).move_to_element(reset_btn).pause(0.4).perform()
                        print("[DEBUG] Наведена мышь на Reset")
                
                        # Ищем снова
                        apply_btn = WebDriverWait(driver, 7).until(
                            EC.element_to_be_clickable((By.XPATH, "//section[@role='dialog']//div[text()='Apply']"))
                        )
                        print("[DEBUG] ✅ Apply найден после реанимации")
                
                    # Прокручиваем к элементу
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", apply_btn)
                    time.sleep(0.5)
                
                    # Проверка перекрытия (через JS)
                    overlapping = driver.execute_script("""
                        const rect = arguments[0].getBoundingClientRect();
                        const x = rect.left + rect.width / 2;
                        const y = rect.top + rect.height / 2;
                        const el = document.elementFromPoint(x, y);
                        return el !== arguments[0] ? el.outerHTML : null;
                    """, apply_btn)
                
                    if overlapping:
                        print(f"[WARNING] Apply может быть перекрыт элементом: {overlapping}")
                    else:
                        print("[DEBUG] ✅ Apply не перекрыт — безопасен для клика")
                
                    # Кликаем
                    try:
                        random_mouse_move(driver)
                        random_scroll(driver)
                        ActionChains(driver).move_to_element(apply_btn).pause(0.3).click().perform()
                        print("[SUCCESS] Клик по Apply выполнен через ActionChains")
                    except Exception as click_exc:
                        print(f"[WARNING] ActionChains не сработал: {click_exc}. Пробуем через JS...")
                        driver.execute_script("arguments[0].click();", apply_btn)
                    
                    # Переключаем график в режим MCap перед анализом сделок
                    print("[SUCCESS] Клик по Apply выполнен через ActionChains")
                    toggle_to_mcap(driver)

                    # Обновляем стейт
                    if 'update_resume_state' in globals():
                        update_resume_state(wallet_index=wallet_index, token_index=i, stage="filter_applied")
                
                    time.sleep(random.uniform(1.2, 2.0))
                
                except Exception as e:
                    print(f"[ERROR] ❌ Не удалось нажать Apply: {e}")
                    driver.save_screenshot("error_apply_click.png")
                    driver.back()
                    solve_turnstile_if_present(driver)
                    time.sleep(random.uniform(3.5, 6.5))  # DElAY * 5 → не фиксирован
                    continue

                # --- Сбор данных по сделкам ---
                print("[DEBUG] Начинаем сбор данных по сделкам...")
                
                try:
                    print("[DEBUG] Начинаем сбор данных по сделкам через BeautifulSoup...")
                    print("[DEBUG] ⏳ Ждём появления контейнера сделок (g-table-body)...")
                
                    container = WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'g-table-body')]"))
                    )
                
                    driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", container)
                    time.sleep(2)
                    print("[DEBUG] ✅ Контейнер сделок найден и прокручен")
                
                    html = container.get_attribute("outerHTML")
                    soup = BeautifulSoup(html, "html.parser")
                    rows = soup.select("div[data-index]")

                    print(f"[DEBUG]  Найдено строк сделок: {len(rows)}")
                    if not rows:
                        print("[SKIP] ❌ Сделки отсутствуют полностью — пропускаем токен")
                        driver.back()
                        solve_turnstile_if_present(driver)
                        time.sleep(random.uniform(3.5, 6.5))
                        continue
                
                    all_buys = []
                    all_sells = []
                
                    for row in rows:
                        try:
                            text = row.get_text(separator="|")
                            cells = text.split("|")
                
                            if len(cells) < 5:
                                print(f"[WARN] Пропущена строка (мало данных): {text}")
                                continue
                
                            action = cells[1].strip().upper()
                            time_str = cells[0].strip()
                            usd = cells[2].strip()
                            mcap = cells[4].strip()
                
                            try:
                                time_obj = datetime.strptime(time_str, "%m/%d %H:%M:%S").replace(year=datetime.utcnow().year)
                            except Exception as e:
                                print(f"[WARN] ⚠️ Не удалось распарсить время: {time_str} — {e}")
                                continue
                
                            record = {
                                "time": time_str,
                                "usd": usd,
                                "mcap": mcap,
                                "dt": time_obj  # добавляем для сортировки
                            }
                
                            print(f"[DEBUG] [{action}] Время: {time_str}, USD: {usd}, MC: {mcap}")
                
                            if action == "BUY":
                                all_buys.append(record)
                            elif action == "SELL":
                                all_sells.append(record)
                
                        except Exception as e:
                            print(f"[WARN] ⚠️ Ошибка обработки строки сделки: {e}")
                
                    # ������ Сортируем и берём первую сделку каждого типа
                    first_buy = sorted(all_buys, key=lambda x: x["dt"])[0] if all_buys else None
                    first_sell = sorted(all_sells, key=lambda x: x["dt"])[0] if all_sells else None
                
                    if not first_buy:
                        print("[SKIP] ❌ Не найдена ни одна покупка — пропускаем токен")
                        driver.back()
                        solve_turnstile_if_present(driver)
                        time.sleep(random.uniform(3.5, 6.5))
                        continue
                    if not first_sell:
                        print("[INFO] Продажа не найдена, заполняем пустыми значениями")
                        first_sell = {"time": "N/A", "usd": "N/A", "mcap": "N/A"}
                
                    print(f"[SUCCESS] ✅ Первая покупка: {first_buy}")
                    print(f"[SUCCESS] ✅ Первая продажа: {first_sell}")

                    try:
                        print("[GRAPH] Возвращаем график в область видимости...")
                    
                        graph_header = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div/div[1]/main/div/div[2]/div[1]"))
                        )
                    
                        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", graph_header)
                        time.sleep(random.uniform(1.6, 2.6))
                    
                        ActionChains(driver).move_to_element(graph_header).pause(random.uniform(0.4, 0.7)).perform()
                        print("[GRAPH] График должен быть виден")
                    
                    except Exception as e:
                        print(f"[GRAPH] Не удалось вернуть график в зону видимости: {e}")
                    
                    # === Сначала обновляем, что сделки разобраны
                    update_resume_state(wallet_index=wallet_index, token_index=i, stage="trades_parsed")
                    
                    # ✅ ДО загрузки графика, фиксируем переход к следующему этапу
                    update_resume_state(wallet_index=wallet_index, token_index=i, stage="before_graph_analysis")

                    # Получаем адрес контракта токена
                    contract_url = token.get("link") or token.get("Token URL")
                    if not contract_url:
                        raise Exception("[CRITICAL] ❌ У токена отсутствует URL!")
                    
                    contract_address = contract_url.split("/")[-1].split("?")[0]
                    print(f"[DEBUG] Контракт токена: {contract_address}")
                    
                    # ⚙ Получаем и дозаписываем историю MCAP
                    mcap_timeline = load_or_update_token_history(driver, contract_address)
                    
                    max_mcap_after_buy = None
                    buy_dt = first_buy["dt"]
                    
                    # фильтруем timeline по дате покупки
                    print(f"[DEBUG] Покупка трейдера: {first_buy['dt']} (тип: {type(first_buy['dt'])})")
                    print(f"[DEBUG] Первый dt в mcap_timeline: {mcap_timeline[0][0]} (тип: {type(mcap_timeline[0][0])})")
                    
                    filtered = []
                    for dt, val in mcap_timeline:
                        candle_start = dt
                        candle_end = dt + timedelta(hours=4)
                        if candle_start <= buy_dt < candle_end:
                            filtered.append((dt, val))
                        elif dt > buy_dt:
                            filtered.append((dt, val))
                    
                    if filtered:
                        max_val_entry = max(filtered, key=lambda x: format_mcap(x[1]))
                        max_mcap_after_buy = format_mcap(max_val_entry[1])
                        max_mcap_after_buy_date = max_val_entry[0]
                        print(f"[INFO] ✅ Max MCap после покупки: {max_mcap_after_buy:.2f} на {max_mcap_after_buy_date.strftime('%Y-%m-%d %H:%M')}")
              
                    update_resume_state(wallet_index=wallet_index, token_index=i, stage="graph_analyzed")
                    
                    # Реальные иксы
                    first_buy_mcap = format_mcap(first_buy["mcap"])
                    real_max_x = round(max_mcap_after_buy / first_buy_mcap, 2) if max_mcap_after_buy and first_buy_mcap else None

                    # добавление в итог
                    single_token = {
                        "Token": token_name,
                        "First Buy Time": first_buy["time"],
                        "First Buy USD": first_buy["usd"],
                        "First Buy Price": first_buy["mcap"],
                        "First Sell Time": first_sell["time"],
                        "First Sell USD": first_sell["usd"],
                        "Real Max X": f"{real_max_x:.2f}x" if real_max_x else "N/A",
                        "First Sell Price": first_sell["mcap"],
                        "Max Marketcap": f"{max_mcap_after_buy:.2f}" if max_mcap_after_buy else "N/A",
                        "Token URL": token_link
                    }
                
                    # === Сохраняем Excel и удаляем временные файлы ===
                    try:
                        save_token_result(wallet, [single_token], balance_sol=balance_sol)
                    
                        # Удаляем лог только при успешном сохранении
                        log_path = "/home/chromeuser/screens/mcap_history/mcap_history_log.txt"
                        if os.path.exists(log_path):
                            os.remove(log_path)
                            print("[CLEANUP] ✅ Удалён лог mcap_history_log.txt после сохранения Excel")
                    
                        # Удаляем PNG-файлы только при успехе
                        for f in glob.glob("/home/chromeuser/screens/mcap_history/*.png"):
                            os.remove(f)
                        print("[CLEANUP] ✅ Удалены PNG-файлы анализа графика после сохранения")
                    
                    except Exception as e:
                        print(f"[ERROR] ❌ Ошибка при сохранении Excel-результата или удалении логов: {e}")
                        driver.save_screenshot("error_save_token_result.png")
                    
                    # === Очистка временных файлов вне зависимости от успеха анализа ===
                    try:
                        for fname in ["ocr_dates_debug.txt", "mcap_history_log.txt"]:
                            fpath = os.path.join("/home/chromeuser/screens/mcap_history", fname)
                            if os.path.exists(fpath):
                                os.remove(fpath)
                                print(f"[CLEANUP] ✅ Удалён файл: {fname}")
                    except Exception as e:
                        print(f"[CLEANUP][ERROR] ❌ Не удалось удалить временные файлы: {e}")
                        
                except Exception as e:
                    print(f"[ERROR] ❌ Ошибка при анализе графика или сохранении отчёта: {e}")
                    driver.save_screenshot("error_save_token_result.png")
                                      
                # Возврат на страницу кошелька через повторный поиск
                random_scroll(driver)
                random_mouse_move(driver)
                navigate_to_wallet_via_search(driver, wallet)
                solve_turnstile_if_present(driver)
                close_pnl_modal_if_present(driver)
                driver.execute_script("window.scrollTo(0, 0);")
                time.sleep(random.uniform(2.5, 4.5))
                
                continue
                
        except Exception as e:
            print(f"[ERROR] ❌ Не удалось получить Recent PnL токены: {e}")
            return results

    except Exception as e:
        print(f"[ERROR] Ошибка с кошельком {wallet}: {e}")

    return results

def format_mcap(raw):
    if isinstance(raw, str):
        raw = raw.replace("$", "").replace(",", "").replace("H", "").strip().upper()
        if "K" in raw:
            return float(raw.replace("K", "")) * 1_000
        elif "M" in raw:
            return float(raw.replace("M", "")) * 1_000_000
        elif "B" in raw:
            return float(raw.replace("B", "")) * 1_000_000_000
        else:
            return float(raw)
    return float(raw)

def human_mcap(num):
    if num >= 1_000_000_000:
        return f"{round(num/1_000_000_000, 2)}b"
    elif num >= 1_000_000:
        return f"{round(num/1_000_000, 2)}m"
    elif num >= 1_000:
        return f"{round(num/1_000, 2)}k"
    else:
        return f"{round(num, 2)}"

def human_duration(td):
    seconds = int(td.total_seconds())
    if seconds < 60:
        duration = f"{int(seconds)} sec"
    elif seconds < 3600:
        duration = f"{int(seconds // 60)} min"
    else:
        h = int(seconds // 3600)
        m = int((seconds % 3600) // 60)
        duration = f"{h}h {m}min"
    return duration

def parse_usd(text):
    """
    Преобразует строку в формат float, очищая $, запятые и пробелы.
    Пример: "$1,234.56" → 1234.56
    """
    try:
        return float(text.replace("$", "").replace(",", "").strip())
    except Exception as e:
        print(f"[WARN] Не удалось распарсить USD: {text} → {e}")
        return 0.0

def get_sol_usd_rate():
    try:
        response = requests.get("https://api.coingecko.com/api/v3/simple/price?ids=solana&vs_currencies=usd")
        data = response.json()
        return float(data["solana"]["usd"])
    except Exception as e:
        print(f"[ERROR] Не удалось получить курс SOL: {e}")
        return None

def save_token_result(wallet, token_data_list, balance_sol=None):
    if not token_data_list:
        print(f"[WARN] Нет данных для отчета по кошельку {wallet}")
        return

    from openpyxl.utils import get_column_letter

    short = wallet[:4]
    report_path = f"./reports/Report_{short}.xlsx"
    os.makedirs("reports", exist_ok=True)

    if os.path.exists(report_path):
        print(f"[INFO] Файл уже существует — дозаписываем: {report_path}")
        existing_wb = openpyxl.load_workbook(report_path)
        existing_ws = existing_wb.active

        # Собираем ключи токенов (Token, First Buy Time)
        existing_keys = set()
        for row in existing_ws.iter_rows(min_row=4, max_row=existing_ws.max_row):
            token = row[0].value
            time = row[7].value if len(row) >= 8 else None
            if token and time:
                existing_keys.add((str(token).strip(), str(time).strip()))
    else:
        existing_wb = openpyxl.Workbook()
        existing_ws = existing_wb.active
        existing_ws.title = "Report"

        # Шапка
        existing_ws["A1"] = "Wallet"
        existing_ws["B1"] = "Balance (SOL)"
        existing_ws["C1"] = "Total Tokens"
        existing_ws["D1"] = "Max Rocket"
        existing_ws["E1"] = "WinRate"
        existing_ws["F1"] = ">5x"
        existing_ws["G1"] = ">10x"

        existing_ws["A2"] = wallet
        existing_ws["B2"] = float(balance_sol) if balance_sol is not None else "—"
        existing_ws["B2"].number_format = "0.000000"

        headers = [
            "Token", "Spent", "Duration to Last", "Mcap at First Buy",
            "Max Token MCap after Buy", "ROI (%)", "Max Profit", "Real Max X",
            "First Trade", "Last Trade", "Token URL"
        ]
        for col, header in enumerate(headers, 1):
            cell = existing_ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DEEAF6", end_color="DEEAF6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        existing_keys = set()

    # Определяем позицию для вставки
    next_row = existing_ws.max_row + 1

    greater_5x = 0
    greater_10x = 0
    total_first_mcap = 0
    total_last_mcap = 0
    durations = []
    
    sol_usd_rate = get_sol_usd_rate()
    if not sol_usd_rate:
        print("[WARN] Курс SOL не получен, пропускаем расчёт в SOL")

    for token in token_data_list:
        key = (token["Token"].strip(), token["First Buy Time"].strip())
        if key in existing_keys:
            print(f"[SKIP] Повтор токена: {token['Token']} в {token['First Buy Time']}")
            continue

        try:
            spent = parse_usd(token["First Buy USD"])
            if sol_usd_rate:
                spent_sol = round(spent / sol_usd_rate, 3)
                spent_str = f"{spent_sol} sol"
            else:
                spent_str = f"{spent} USD"
            mcap_buy = format_mcap(token["First Buy Price"])
            mcap_max = format_mcap(token["Max Marketcap"]) if token["Max Marketcap"] != "N/A" else None
            mcap_sell = format_mcap(token["First Sell Price"]) if token["First Sell Price"] != "N/A" else None

            total_first_mcap += mcap_buy
            if mcap_sell:
                total_last_mcap += mcap_sell

            # ROI
            roi = None
            if mcap_sell and mcap_sell > 0:
                roi = ((mcap_sell - mcap_buy) / mcap_buy) * 100

            # Max Profit
            max_profit_x = None
            if mcap_max and mcap_buy:
                max_profit_x = mcap_max / mcap_buy
                if max_profit_x >= 5:
                    greater_5x += 1
                if max_profit_x >= 10:
                    greater_10x += 1

            # Duration
            duration = ""
            try:
                if token["First Buy Time"] != "N/A" and token["First Sell Time"] != "N/A":
                    dt1 = datetime.strptime(token["First Buy Time"], "%m/%d %H:%M:%S")
                    dt2 = datetime.strptime(token["First Sell Time"], "%m/%d %H:%M:%S")
                    delta = dt2 - dt1
                    durations.append(delta)
                    seconds = delta.total_seconds()
                    if seconds < 60:
                        duration = f"{int(seconds)} sec"
                    elif seconds < 3600:
                        duration = f"{int(seconds // 60)} min"
                    else:
                        h = int(seconds // 3600)
                        m = int((seconds % 3600) // 60)
                        duration = f"{h}h {m}min"
            except:
                duration = "N/A"

            values = [
                token["Token"],
                spent_str,
                duration,
                human_mcap(mcap_buy),
                human_mcap(mcap_max) if mcap_max else "N/A",
                f"{round(roi, 2)}%" if roi is not None else "N/A",
                f"{round(max_profit_x, 2)}x" if max_profit_x else "N/A",
                token.get("Real Max X", "N/A"),
                token["First Buy Time"],
                token["First Sell Time"],
                token["Token URL"]
            ]

            for col, val in enumerate(values, 1):
                cell = existing_ws.cell(row=next_row, column=col, value=val)
                cell.alignment = Alignment(horizontal="center")

                if col == 6 and roi is not None:
                    cell.fill = PatternFill(
                        start_color="C6EFCE" if roi >= 0 else "FFC7CE",
                        end_color="C6EFCE" if roi >= 0 else "FFC7CE",
                        fill_type="solid"
                    )

                if col == 7 and max_profit_x is not None:
                    val_float = float(str(val).replace("x", "").strip())
                    cell.fill = PatternFill(
                        start_color="C6EFCE" if val_float >= 1 else "FFC7CE",
                        end_color="C6EFCE" if val_float >= 1 else "FFC7CE",
                        fill_type="solid"
                    )

            # Гиперссылка
            contract_address = token["Token URL"].split("/")[-1].split("?")[0]
            full_url = f"https://gmgn.ai/sol/token/{contract_address}?maker={wallet}"
            link_cell = existing_ws.cell(row=next_row, column=10)
            link_cell.value = "gmgn"
            link_cell.hyperlink = full_url
            link_cell.font = Font(color="0000FF", underline="single")
            link_cell.alignment = Alignment(horizontal="center")

            next_row += 1  # переходим на следующую строку
            existing_keys.add(key)  # чтобы даже в этом же вызове не было дубликатов

        except Exception as e:
            print(f"[ERROR] Ошибка при записи токена в Excel: {e}")

    # Обновляем верхние метрики с учётом Max Rocket и WinRate
    existing_ws["A2"] = wallet
    existing_ws["B2"] = float(balance_sol) if balance_sol is not None else "N/A"
    existing_ws["B2"].number_format = '0.000'
    existing_ws["C2"] = len(existing_keys)
    existing_ws["F2"] = greater_5x
    existing_ws["G2"] = greater_10x
    
    # Max Rocket (из колонки G / индекс 6)
    max_rocket_val = 0
    for row in existing_ws.iter_rows(min_row=4, max_row=existing_ws.max_row):
        val = row[6].value
        if val and isinstance(val, str) and "x" in val:
            try:
                x_val = float(val.replace("x", "").strip())
                if x_val > max_rocket_val:
                    max_rocket_val = x_val
            except:
                pass
    existing_ws["D1"] = "Max Rocket"
    existing_ws["D2"] = f"{max_rocket_val}x"
    
    # WinRate (по колонке ROI — индекс 5)
    total_roi = 0
    positive_roi = 0
    for row in existing_ws.iter_rows(min_row=4, max_row=existing_ws.max_row):
        roi_cell = row[5].value
        if roi_cell and isinstance(roi_cell, str) and "%" in roi_cell:
            try:
                roi_val = float(roi_cell.replace("%", "").strip())
                total_roi += 1
                if roi_val > 0:
                    positive_roi += 1
            except:
                pass
    existing_ws["E1"] = "WinRate"
    existing_ws["E2"] = f"{round((positive_roi / total_roi) * 100, 1)}%" if total_roi else "N/A"
    
    # === Пересчёт реальных значений >5x и >10x по всей таблице ===
    greater_5x = 0
    greater_10x = 0
    for row in existing_ws.iter_rows(min_row=4, max_row=existing_ws.max_row):
        val = row[6].value  # колонка G, "Max Profit"
        if val and isinstance(val, str) and "x" in val:
            try:
                x_val = float(val.replace("x", "").strip())
                if x_val >= 5:
                    greater_5x += 1
                if x_val >= 10:
                    greater_10x += 1
            except:
                continue
    
    existing_ws["F2"] = greater_5x
    existing_ws["G2"] = greater_10x

    # Заголовки и выравнивание
    existing_ws["A1"] = "Wallet"
    existing_ws["B1"] = "Balance (SOL)"
    existing_ws["C1"] = "Total Tokens"
    existing_ws["F1"] = ">5x"
    existing_ws["G1"] = ">10x"
    for col in range(1, 8):
        existing_ws.cell(row=1, column=col).font = Font(bold=True)
        existing_ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")
        existing_ws.cell(row=2, column=col).alignment = Alignment(horizontal="center")

    # Автоширина
    for col in range(1, 11):
        max_len = 0
        for row in existing_ws.iter_rows(min_row=1, max_row=existing_ws.max_row, min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        existing_ws.column_dimensions[get_column_letter(col)].width = max_len + 2

    existing_wb.save(report_path)
    print(f"[SAVED] ✅ Отчет сохранен: {report_path}")

def load_or_update_token_history(driver, contract_address):
    history_dir = "/home/chromeuser/screens/token_cache"
    os.makedirs(history_dir, exist_ok=True)

    filepath = os.path.join(history_dir, f"{contract_address}.json")

    existing = {}
    if os.path.exists(filepath):
        try:
            with open(filepath, "r") as f:
                existing = json.load(f)
        except Exception as e:
            print(f"[HISTORY] ⚠ Ошибка чтения истории токена: {e}")

    # Конвертируем в datetime
    existing_parsed = {
        datetime.strptime(k, "%Y-%m-%d %H:%M"): v
        for k, v in existing.items()
    } if existing else {}

    if existing_parsed:
        last_known = max(existing_parsed.keys())
        now = datetime.utcnow()
        if now - last_known < timedelta(days=1):
            print(f"[HISTORY]  История {contract_address} актуальна, используем кэш")
            return sorted(existing_parsed.items())

    print(f"[HISTORY]  Обновляем историю токена {contract_address}...")

    # Запрашиваем новые данные
    timeline_new = run_detailed_mcap_scraper(driver, contract_address=contract_address)

    combined = {**{dt.strftime("%Y-%m-%d %H:%M"): val for dt, val in existing_parsed},
                **{dt.strftime("%Y-%m-%d %H:%M"): val for dt, val in timeline_new}}

    try:
        with open(filepath, "w") as f:
            json.dump(combined, f, indent=2)
        print(f"[HISTORY] ✅ История токена сохранена: {filepath}")
    except Exception as e:
        print(f"[HISTORY] ❌ Ошибка сохранения истории токена: {e}")

    result = []
    for k, v in combined.items():
        if isinstance(k, str) and isinstance(v, str):
            try:
                dt = datetime.strptime(k, "%Y-%m-%d %H:%M")
                result.append((dt, v))
            except Exception as e:
                print(f"[HISTORY] ⚠ Проблема с элементом истории: {k} — {e}")
    return sorted(result)

def main():
    # ✅ Загружаем сохранённое состояние
    state = load_resume_state()
    resume_wallet_index = state["wallet_index"]
    resume_token_index = state["token_index"]
    resume_stage = state["stage"]

    try:
        # Загрузка Excel с адресами кошельков
        wallet_files = [f for f in os.listdir(DOWNLOAD_FOLDER) if f.endswith(('.xlsx', '.xls', '.csv'))]
        if not wallet_files:
            print("[ERROR] Не найден файл с кошельками в папке downloads")
            return

        wallet_file = wallet_files[0]
        try:
            if wallet_file.endswith('.csv'):
                df_wallets = pd.read_csv(os.path.join(DOWNLOAD_FOLDER, wallet_file))
            else:
                df_wallets = pd.read_excel(os.path.join(DOWNLOAD_FOLDER, wallet_file))

            wallets = df_wallets.iloc[:, 0].dropna().astype(str).tolist()
            print(f"[INFO] Загружено кошельков: {len(wallets)}")
        except Exception as e:
            print(f"[ERROR] Ошибка при загрузке файла с кошельками: {str(e)}")
            return

        # Подключение к Chrome
        chrome_options = Options()
        chrome_options.debugger_address = "localhost:9222"
        driver = webdriver.Chrome(options=chrome_options)
        driver.maximize_window()

        all_results = []

        try:
            for i, wallet in enumerate(wallets):
                if i < resume_wallet_index:
                    continue  # Пропускаем ранее обработанные

                print(f"\n[PROGRESS] Обработка кошелька {i+1}/{len(wallets)}")

                # Только если это НЕ восстановление — логируем с нуля
                if i != resume_wallet_index:
                    save_resume_state(i, 0, "start_wallet")

                results = process_wallet(
                    driver,
                    wallet,
                    wallet_index=i,
                    token_resume_index=resume_token_index if i == resume_wallet_index else 0,
                    resume_stage=resume_stage if i == resume_wallet_index else "start"
                )
                all_results.extend(results)

                # ✅ Удаляем файл состояния только если обработка кошелька успешна
                if results:
                    clear_resume_state()

                # Промежуточное сохранение
                if (i + 1) % 5 == 0 and all_results:
                    now = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                    report_path = os.path.join(REPORT_FOLDER, f"Wallet_Analysis_Report_{now}_part_{i//5 + 1}.xlsx")
                    pd.DataFrame(all_results).to_excel(report_path, index=False)
                    print(f"[INFO] Промежуточный отчет сохранен: {report_path}")

        except KeyboardInterrupt:
            print("\n[INTERRUPT] Скрипт остановлен вручную — сохраняем прогресс и выходим")
            # ⚠ Файл состояния сохраняется, но НЕ удаляется

        # Финальное сохранение отчета
        if all_results:
            now = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            report_path = os.path.join(REPORT_FOLDER, f"Wallet_Analysis_Report_{now}_final.xlsx")
            pd.DataFrame(all_results).to_excel(report_path, index=False)
            print(f"\n[SUCCESS] Финальный отчет сохранен: {report_path}")
            print(f"[INFO] Всего обработано кошельков: {len(wallets)}")
            print(f"[INFO] Всего собрано записей: {len(all_results)}")
        else:
            print("\n[WARNING] Не удалось собрать данные")

    except Exception as e:
        print(f"\n[CRITICAL] Основная ошибка: {str(e)}")
    finally:
        if 'driver' in locals():
            driver.quit()

if __name__ == "__main__":
    main()